import os
from flask import Flask, render_template, request
import openpyxl

app = Flask(__name__)


# دالة لقراءة النتيجة من ملف الإكسيل
def get_student_result(seat_no):
    excel_path = os.path.join(os.path.dirname(__file__), 'results.xlsx')
    if not os.path.exists(excel_path):
        return None

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active

    # البحث عن رقم الجلوس في العمود الأول (A)
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value and str(cell_value).strip() == str(seat_no).strip():
            # قراءة العناوين (المواد) من الصف الأول
            headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
            # قراءة بيانات الطالب من الصف الحالي
            student_data = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]

            result = {}
            for header, data in zip(headers, student_data):
                if header:
                    if 'اسم' in str(header):
                        result['student_name'] = data
                    elif 'جلوس' in str(header) or 'رقم' in str(header) and len(result) == 0:
                        result['seat_no'] = data
                    elif 'مجموع' in str(header) or 'gpa' in str(header).lower() or 'تقدير عام' in str(header).lower():
                        result['gpa'] = data
                    elif 'حالة' in str(header) or 'النتيجة' in str(header):
                        result['status'] = data
                    else:
                        result[header] = data

            # التأكد من وجود المفاتيح الأساسية لتجنب أخطاء الـ HTML
            if 'seat_no' not in result: result['seat_no'] = seat_no
            if 'student_name' not in result: result['student_name'] = "طالب"
            if 'gpa' not in result: result['gpa'] = "غير محدد"
            if 'status' not in result: result['status'] = "ن"

            return result
    return None


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        seat_no = request.form.get('seat_no')
        result = get_student_result(seat_no)
        if result:
            # تحويل أي رموز في النتيجة لنصوص نضيفة عشان الـ HTML يقراها علطول بدون إيرور str
            cleaned_result = {}
            for k, v in result.items():
                cleaned_result[k] = str(v).strip()
            return render_template('index.html', searched=True, result=cleaned_result)
        else:
            return render_template('index.html', searched=True, result=None,
                                   error_msg="عفواً، رقم الجلوس هذا غير مسجل أو غير صحيح.")
    return render_template('index.html', searched=False)


if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=True)